import pandas as pd
import openpyxl
import os
import shutil 
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from collections import defaultdict

# Načtení Excelu do pandas DataFrame s pomocí openpyxl pro práci se styly
file_path = 'downloadedFlights.xlsx'
wb = load_workbook(filename=file_path, data_only=True)
ws = wb["PED_LŘ"]

# Extrakce hodnot a stylů z Excelu
data = []
for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    # Získání fontu a jeho vlastností (barva, přeškrtnutí)
    font = ws.cell(row=i, column=1).font
    color = font.color.rgb if font.color else None
    strike = font.strike

    # Ignorujeme přeškrtnuté nebo červené hodnoty
    if strike or (color and color.lower() == 'ff0000'):  # Červená barva v hex formátu #FF0000
        continue

    data.append([cell.value for cell in row])

# Vytvoření DataFrame z filtrovaných dat
column_names = [
    'Index', 'Day', 'Status', 'Flight', 'Departure_Time', 'Arrival_Time', 
    'Airline', 'Departure_City', 'Arrival_City', 'Aircraft_Type', 
    'Unknown_Column_1', 'Delay', 'Unknown_Column_2'
]

# Přiřaď názvy sloupců do DataFrame
df = pd.DataFrame(data)
column_names = ['D', 'M', 'DEN', 'ČÍSLO LETU', 'PŘÍLET', 'ODLET', 'SPOLEČNOST', 'PŘÍLET Z',	'ODLET DO',	'Typ A/C',	'POZNÁMKY',	'ODB.', 'None'		
]

# Přiřaď názvy sloupců do DataFrame
df.columns = column_names

# Odstranění tečky z konce čísel ve sloupcích D a M
df['D'] = df['D'].astype(str).str.replace(r'\.$', '', regex=True)
df['M'] = df['M'].astype(str).str.replace(r'\.$', '', regex=True)

# Získání aktuálního data
current_date = datetime.now()

# Převod na datum s vyčištěním hodnot
df['Datum'] = pd.to_datetime(df['D'] + '-' + df['M'] + '-' + str(current_date.year), format='%d-%m-%Y', errors='coerce')

# Odstranění řádků s neplatným datem
df = df.dropna(subset=['Datum'])


# Filtrování podle aktuálního data a maximálně do pondělí aktuálního týdne
start_of_week = current_date - timedelta(days=current_date.weekday() + 1)  # Začátek týdne (pondělí)
end_date = start_of_week + timedelta(days=14)
df_filtered = df[(df['Datum'] <= end_date) & (df['Datum'] >= start_of_week)]

# Extrakce kódu města ze sloupce 'ODLET DO'
df_filtered['Město Kód'] = df_filtered['ODLET DO'].str.split('-').str[-1].str.strip()

# Výsledek
print(df_filtered[['Datum', 'ODLET', 'Město Kód', 'ODB.']])

# Uložení zpracovaných dat do pole (listu) pro další práci
data_list = df_filtered[['Datum', 'ODLET', 'Město Kód', 'ODB.']].values.tolist()



# Cesta k šabloně
template_path = 'template.xlsx'
output_path = 'rozpis.xlsx'

if os.path.exists(output_path):
    os.remove(output_path)  # Pokud soubor existuje, smaže ho
    
shutil.copyfile(template_path,output_path) #Kopírování šablony

wb = load_workbook(filename=output_path)
list1 = wb["List1"]
list2 = wb["List2"]


# Data, která chceme zpracovat (předpoklad: list tuple nebo DataFrame)
# data_list = [(datum, odlet, mesto, odb)]  <- formát dat

# Rozdělení dat podle jednotlivých dnů
data_by_day = defaultdict(list)

for datum, odlet, mesto, odb in data_list:
    data_by_day[datum.date()].append((odlet, mesto, odb))

# Seřazení dat podle dnů
sorted_days = sorted(data_by_day.keys())

# Funkce pro vyplnění listu
def fill_sheet(sheet, sorted_days, data_by_day, start_index=0):
    for i in range(7):
        if start_index + i >= len(sorted_days):
            break

        current_day = sorted_days[start_index + i]
        flights = data_by_day[current_day]
        
        try:
            # Získání textového obsahu pojmenovaného rozsahu
            date_cell_range = sheet.defined_names['date' + str(i + 1)].attr_text
            odlet_cell_range = sheet.defined_names['check' + str(i + 1)].attr_text
            odb_cell_range = sheet.defined_names['deps' + str(i + 1)].attr_text
        except KeyError:
            raise ValueError(f"Pojmenovaný rozsah 'date{i+1}' nebo jiný neexistuje ve vašem listu")

        # Extrahování samotného odkazu na buňku z textu
        date_cell = date_cell_range.split('!')[-1]
        odlet_cell = odlet_cell_range.split('!')[-1]
        odb_cell = odb_cell_range.split('!')[-1]

        # Naplnění buněk
        sheet[date_cell].value = current_day.strftime('%d.%m.%Y')

        # Časy odletů beze změny
        odlet_values = ' '.join(
            [f"{f[0].strftime('%H:%M') if hasattr(f[0], 'strftime') else f[0]}({f[1]})" for f in flights]
        )

        # Čas ODB s kódem města v závorce
        odb_values = ' '.join(
            [f"{f[2].strftime('%H:%M') if hasattr(f[2], 'strftime') else f[2]}" for f in flights]
        )

        # Zápis hodnot do buněk
        sheet[odlet_cell].value = odlet_values
        sheet[odb_cell].value = odb_values
        
# Vyplnění týdne
fill_sheet(list1, sorted_days, data_by_day, start_index=0)
fill_sheet(list2, sorted_days, data_by_day, start_index=7)


# Funkce pro vyplnění směn do excelového listu s použitím pojmenovaných rozsahů
def fill_shifts_in_sheet(sheet, shifts, kitchen_cells):
    for day_index, shift_times in enumerate(shifts):
        try:
            # Získání pojmenovaného rozsahu z Excelu
            kitchen_cell_range = sheet.defined_names[kitchen_cells[day_index]].attr_text
            
            # Extrahování samotné buňky z textu pojmenovaného rozsahu
            # Ujistíme se, že extrahujeme pouze první buňku (ne rozsah)
            kitchen_cell = kitchen_cell_range.split('!')[-1].split(':')[0]  # Bereme pouze první buňku, pokud jde o rozsah
        except KeyError:
            raise ValueError(f"Pojmenovaný rozsah '{kitchen_cells[day_index]}' neexistuje ve vašem listu")

        # Vytvoříme text směn, každý na nový řádek
        shifts_text = '\n'.join([f"{shift[0].strftime('%H:%M')} ({shift[2]})" for shift in shift_times])
        
        # Vyplníme buňku v Excelu
        sheet[kitchen_cell].value = shifts_text

# Uložení do nového souboru
wb.save(output_path)

print(f'Data byla úspěšně vyplněna do souboru {output_path}.')